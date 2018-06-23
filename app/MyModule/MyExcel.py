"""
This class is used to generate the report
"""


class MyExcel:
    def __init__(self, **kwargs):
        try:
            self.kwargs = kwargs
            self.dest_file = self.kwargs['dest_file']
            # c_l means create or load workbook
            self.c_l = kwargs.setdefault('c_l', 'c')

            self.data = kwargs.setdefault('data', {})
            self.col_start = kwargs.setdefault('col_start', 1)
            self.row_start = kwargs.setdefault('row_start', 1)
        except KeyError:
            print('destination file does not defined')
            exit()

    def open_workbook(self):
        """
        Self defined function to create (not exist) or load a file (exist)
        Returns: workbook handle

        """
        from openpyxl import Workbook
        from openpyxl import load_workbook
        if self.c_l == 'c':
            # judge existence
            return Workbook()

        else:
            try:
                return load_workbook(filename=self.dest_file)
            except:
                print('ERROR! {} cannot be loaded'.format(self.dest_file))

    def write_data(self, worksheet, *style):
        """

        Args:
            worksheet:
            *style: from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

        Returns:
            the end row and column number

        """

        if style:
            pass

        for col_num in self.data:
            col_start = self.col_start
            for row_num in col_num:
                write_cell = worksheet.cell(row=self.row_start, column=col_start)
                write_cell.value = row_num
                col_start += 1
            self.row_start += 1

        return self.row_start, col_start

    def save_excel(self,workbook):
        workbook.save(filename=self.dest_file)

    def save_virtual_excel(self, workbook):
        from openpyxl.writer.excel import save_virtual_workbook
        return save_virtual_workbook(workbook)

if __name__ == '__main__':
    rows = [
        ['Number', 'Batch 1', 'Batch 2'],
        [2, 40, 30, 123123, '13123'],
        [3, 40, 25],
        [4, 50, 30],
        [5, 30, 10],
        [6, 25, 5],
        [7, 50, 10],
    ]

    init_info = {'dest_file':'my_test_01.xlsx',
                 'c_l':'c',
                 'data':rows,
                 'col_start':1,
                 'row_start':1}
    we = MyExcel(**init_info)
    wb = we.open_workbook()
    ws = wb.active
    ws.title = 'test111'
    we.write_data(ws)
    we.save_excel(wb)
    print('finished')